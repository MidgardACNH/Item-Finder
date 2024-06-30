import discord
from discord.ext import commands
import openpyxl

# Replace 'YOUR_BOT_TOKEN' with your actual bot token
BOT_TOKEN = 'BOT_TOKEN'

# Define intents
intents = discord.Intents.default()
intents.message_content = True  # Enable the intent to receive message content

# Create an instance of the bot with intents
bot = commands.Bot(command_prefix='!', intents=intents)

# Function to load items from all sheets in Excel file into dictionary
def load_items_from_excel(file_path):
    items = {}
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 3:
                    item_name = str(row[0]).strip().lower()
                    if item_name not in items:
                        items[item_name] = {
                            'description': str(row[1]).strip(),
                            'sheet_name': sheet_name
                        }
    
    except Exception as e:
        print(f"Error loading Excel file: {e}")
    
    return items

# Path to your Excel file
EXCEL_FILE_PATH = 'C:\path\to\directory\items.xlsx'

# Load items from Excel into dictionary
treasure_islands = load_items_from_excel(EXCEL_FILE_PATH)

# Example: Printing the loaded dictionary
print(treasure_islands)

# Define the find command with chunking
@bot.command(name='find')
async def find(ctx, *, query: str):
    query_lower = query.lower()  # Convert query to lowercase for case insensitivity
    
    found_items = []
    for item_name, item_info in treasure_islands.items():
        if query_lower in item_name:
            sheet_name = item_info['sheet_name']
            found_items.append(f"**Item Name:** {item_name} - **Located on:** {sheet_name}")
    
    if found_items:
        # Chunk the found items into messages of max 2000 characters
        chunks = []
        current_chunk = ""
        for item in found_items:
            if len(current_chunk) + len(item) + 1 > 2000:  # +1 for newline character
                chunks.append(current_chunk)
                current_chunk = ""
            current_chunk += f"{item}\n"
        if current_chunk:
            chunks.append(current_chunk)
        
        # Send each chunk as a separate message
        for chunk in chunks:
            try:
                await ctx.send(f"Items found:\n{chunk}")
            except discord.errors.HTTPException as e:
                await ctx.send(f"Error: {e}")  # Handle the HTTPException if needed
    else:
        await ctx.send('Item not found. Please try another name.')

# Define the on_ready event
@bot.event
async def on_ready():
    print(f'Bot is ready. Logged in as {bot.user.name}')

# Run the bot
bot.run(BOT_TOKEN)
